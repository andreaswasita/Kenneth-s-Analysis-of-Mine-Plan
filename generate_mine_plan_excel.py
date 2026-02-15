"""
Greenbushes LoM Key Metrics - Data Extraction from Figure 14 (Page 28)
CY25 ORE - IGO Limited Annual Report

Extracts data from all 4 graphs and creates a comprehensive Excel workbook.
Data is for even-numbered years as annotated in the charts.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ==============================================================================
# DATA EXTRACTED FROM FIGURE 14
# ==============================================================================

# Years with annotations (biannual from 2026, even-numbered years)
mining_years = [2026, 2028, 2030, 2032, 2034, 2036, 2038, 2040, 2042, 2044, 2046, 2048]
full_years   = [2026, 2028, 2030, 2032, 2034, 2036, 2038, 2040, 2042, 2044, 2046, 2048, 2050, 2052]

# ── GRAPH 1: Mining (Mt) ──
# Strip ratios are annotated directly on the chart
strip_ratios = [3.2, 3.1, 4.3, 5.8, 11.5, 5.0, 4.5, 5.5, 3.4, 1.9, 1.4, 0.8]

# Total movement re-read from chart against y-axis gridlines (0, 20, 40, 60, 80 Mt)
# 2026=42, 2028=42, 2030=50, 2032=60, 2034=75(peak), 2036=50, 2038=55, 2040=45, 2042=31, 2044=18, 2046=12, 2048=8
total_movement = [42.0, 42.0, 50.0, 60.0, 75.0, 50.0, 55.0, 45.0, 31.0, 18.0, 12.0, 8.0]

# Ore and Waste calculated from: Total = Ore * (1 + Strip Ratio)
ore_mt  = [round(t / (1 + sr), 1) for t, sr in zip(total_movement, strip_ratios)]
waste_mt = [round(t - o, 1) for t, o in zip(total_movement, ore_mt)]

# ── GRAPH 2: Stockpiles (Mt) ──
# Values annotated directly on graph
stockpiles = [8.2, 17.6, 21.1, 23.4, 17.0, 16.7, 20.7, 16.2, 14.6, 14.2, 11.2, 9.5, 2.7, 0.5]

# ── GRAPH 3: Processing (Mt) ──
# Mean lithia feed grade (%Li2O) annotated on chart
li2o_grade = [2.11, 2.12, 2.22, 2.04, 2.18, 2.04, 1.97, 1.95, 1.95, 2.03, 1.95, 2.21, 1.19, 1.79]

# Plant-by-plant processing estimated from stacked area chart (y-axis 0-12 Mt)
proc_cgp3 = [0.3, 1.0, 1.2, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 0.3, 0.0]
proc_cgp2 = [2.7, 2.5, 3.0, 3.0, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 3.0, 1.2, 0.2]
proc_cgp1 = [2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 1.0, 0.3]
proc_tgp  = [2.0, 2.5, 1.8, 1.5, 1.5, 2.0, 2.0, 2.0, 1.5, 1.5, 1.5, 1.5, 1.0, 0.5]
proc_total = [round(a+b+c+d, 1) for a,b,c,d in zip(proc_cgp3, proc_cgp2, proc_cgp1, proc_tgp)]

# ── GRAPH 4: Concentrate (Mt) ──
# Total values annotated directly on chart
conc_total_annotated = [1.5, 2.0, 2.0, 1.7, 1.8, 1.7, 1.6, 1.6, 1.6, 1.7, 1.6, 1.9, 0.4, 0.1]

# Plant-by-plant concentrate estimated from stacked area chart (y-axis 0-3.0 Mt)
conc_cgp3 = [0.05, 0.20, 0.25, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.05, 0.00]
conc_cgp2 = [0.45, 0.60, 0.65, 0.50, 0.50, 0.50, 0.50, 0.50, 0.50, 0.50, 0.50, 0.70, 0.15, 0.03]
conc_cgp1 = [0.50, 0.50, 0.50, 0.40, 0.50, 0.40, 0.40, 0.40, 0.40, 0.50, 0.40, 0.50, 0.12, 0.04]
conc_tgp  = [0.50, 0.70, 0.60, 0.50, 0.50, 0.50, 0.40, 0.40, 0.40, 0.40, 0.40, 0.40, 0.08, 0.03]

# ==============================================================================
# STYLING
# ==============================================================================

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
DATA_FONT = Font(name="Calibri", size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
NOTE_FONT = Font(name="Calibri", italic=True, size=10, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, is_total_row=False):
    cell = ws.cell(row=row, column=col)
    cell.font = TOTAL_FONT if is_total_row else DATA_FONT
    if is_total_row:
        cell.fill = TOTAL_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def auto_width(ws, max_col, min_width=12):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, 15)


# ==============================================================================
# CREATE WORKBOOK
# ==============================================================================

wb = Workbook()

# ──────────────────────────────────────────────────────────────
# SHEET 1: Mining
# ──────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Mining"

ws1.merge_cells("A1:G1")
ws1["A1"] = "MINING (Mt) - Greenbushes LoM Schedule (CY25 ORE)"
ws1["A1"].font = TITLE_FONT

ws1.merge_cells("A2:G2")
ws1["A2"] = "Strip Ratio = Waste / Ore. Mining ends in 2049. Even-year data from Figure 14, Page 28."
ws1["A2"].font = NOTE_FONT

headers = ["Year", "Ore (Mt)", "Waste (Mt)", "Total Movement (Mt)", "Strip Ratio", "Ore %", "Waste %"]
for col, h in enumerate(headers, 1):
    ws1.cell(row=4, column=col, value=h)
style_header_row(ws1, 4, len(headers))

for i, year in enumerate(mining_years):
    row = 5 + i
    ore_pct = round(ore_mt[i] / total_movement[i] * 100, 1)
    waste_pct = round(waste_mt[i] / total_movement[i] * 100, 1)
    values = [year, ore_mt[i], waste_mt[i], total_movement[i], strip_ratios[i], ore_pct, waste_pct]
    for col, v in enumerate(values, 1):
        ws1.cell(row=row, column=col, value=v)
        style_data_cell(ws1, row, col)

# Totals row
total_row = 5 + len(mining_years)
ws1.cell(row=total_row, column=1, value="LoM TOTAL")
ws1.cell(row=total_row, column=2, value=round(sum(ore_mt), 1))
ws1.cell(row=total_row, column=3, value=round(sum(waste_mt), 1))
ws1.cell(row=total_row, column=4, value=round(sum(total_movement), 1))
ws1.cell(row=total_row, column=5, value=round(sum(waste_mt) / sum(ore_mt), 1))
ws1.cell(row=total_row, column=6, value=round(sum(ore_mt) / sum(total_movement) * 100, 1))
ws1.cell(row=total_row, column=7, value=round(sum(waste_mt) / sum(total_movement) * 100, 1))
for col in range(1, len(headers) + 1):
    style_data_cell(ws1, total_row, col, is_total_row=True)

# Notes
note_row = total_row + 2
ws1.merge_cells(f"A{note_row}:G{note_row}")
ws1[f"A{note_row}"] = "Notes: Values are for even-numbered years (biannual snapshots). Ore & Waste derived from total movement and strip ratio."
ws1[f"A{note_row}"].font = NOTE_FONT
ws1.merge_cells(f"A{note_row+1}:G{note_row+1}")
ws1[f"A{note_row+1}"] = "Total movement estimated from chart. Strip ratios are directly annotated on the graph. Mining ceases at end of 2049."
ws1[f"A{note_row+1}"].font = NOTE_FONT

auto_width(ws1, len(headers))

# ──────────────────────────────────────────────────────────────
# SHEET 2: Stockpiles
# ──────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Stockpiles")

ws2.merge_cells("A1:D1")
ws2["A1"] = "STOCKPILES (Mt) - Greenbushes LoM Schedule (CY25 ORE)"
ws2["A1"].font = TITLE_FONT

ws2.merge_cells("A2:D2")
ws2["A2"] = "Stockpiles (mostly off-RoM). Values annotated directly on Figure 14. Stockpiles exhausted by ~2053."
ws2["A2"].font = NOTE_FONT

headers2 = ["Year", "Stockpile (Mt)", "Change vs Prior (Mt)", "Change (%)"]
for col, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=col, value=h)
style_header_row(ws2, 4, len(headers2))

for i, year in enumerate(full_years):
    row = 5 + i
    change = round(stockpiles[i] - stockpiles[i-1], 1) if i > 0 else 0
    pct = round(change / stockpiles[i-1] * 100, 1) if i > 0 and stockpiles[i-1] != 0 else 0
    values = [year, stockpiles[i], change, pct]
    for col, v in enumerate(values, 1):
        ws2.cell(row=row, column=col, value=v)
        style_data_cell(ws2, row, col)

# Peak stockpile info
peak_row = 5 + len(full_years) + 1
ws2.merge_cells(f"A{peak_row}:D{peak_row}")
peak_val = max(stockpiles)
peak_yr = full_years[stockpiles.index(peak_val)]
ws2[f"A{peak_row}"] = f"Peak Stockpile: {peak_val} Mt in {peak_yr}"
ws2[f"A{peak_row}"].font = SUBTITLE_FONT

auto_width(ws2, len(headers2))

# ──────────────────────────────────────────────────────────────
# SHEET 3: Processing
# ──────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Processing")

ws3.merge_cells("A1:H1")
ws3["A1"] = "PROCESSING (Mt) BY PLANT - Greenbushes LoM Schedule (CY25 ORE)"
ws3["A1"].font = TITLE_FONT

ws3.merge_cells("A2:H2")
ws3["A2"] = "Plants: TGP, CGP1, CGP2, CGP3. Processing continues through 2053 to exhaust stockpiles."
ws3["A2"].font = NOTE_FONT

headers3 = ["Year", "CGP3 (Mt)", "CGP2 (Mt)", "CGP1 (Mt)", "TGP (Mt)", "Total Processing (Mt)", "Li2O Grade (%)", "Notes"]
for col, h in enumerate(headers3, 1):
    ws3.cell(row=4, column=col, value=h)
style_header_row(ws3, 4, len(headers3))

for i, year in enumerate(full_years):
    row = 5 + i
    note = "Post-mining (stockpile feed)" if year >= 2050 else ""
    values = [year, proc_cgp3[i], proc_cgp2[i], proc_cgp1[i], proc_tgp[i], proc_total[i], li2o_grade[i], note]
    for col, v in enumerate(values, 1):
        ws3.cell(row=row, column=col, value=v)
        style_data_cell(ws3, row, col)

# Totals
total_row3 = 5 + len(full_years)
ws3.cell(row=total_row3, column=1, value="LoM TOTAL")
ws3.cell(row=total_row3, column=2, value=round(sum(proc_cgp3), 1))
ws3.cell(row=total_row3, column=3, value=round(sum(proc_cgp2), 1))
ws3.cell(row=total_row3, column=4, value=round(sum(proc_cgp1), 1))
ws3.cell(row=total_row3, column=5, value=round(sum(proc_tgp), 1))
ws3.cell(row=total_row3, column=6, value=round(sum(proc_total), 1))
ws3.cell(row=total_row3, column=7, value=round(sum(li2o_grade) / len(li2o_grade), 2))
ws3.cell(row=total_row3, column=8, value="Avg Grade")
for col in range(1, len(headers3) + 1):
    style_data_cell(ws3, total_row3, col, is_total_row=True)

auto_width(ws3, len(headers3))

# ──────────────────────────────────────────────────────────────
# SHEET 4: Concentrate
# ──────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Concentrate")

ws4.merge_cells("A1:G1")
ws4["A1"] = "CONCENTRATE (Mt) BY PLANT - Greenbushes LoM Schedule (CY25 ORE)"
ws4["A1"].font = TITLE_FONT

ws4.merge_cells("A2:G2")
ws4["A2"] = "Concentrate produced by each processing plant. Total values annotated on Figure 14."
ws4["A2"].font = NOTE_FONT

headers4 = ["Year", "CGP3 (Mt)", "CGP2 (Mt)", "CGP1 (Mt)", "TGP (Mt)", "Total Concentrate (Mt)", "Recovery Proxy (%)"]
for col, h in enumerate(headers4, 1):
    ws4.cell(row=4, column=col, value=h)
style_header_row(ws4, 4, len(headers4))

for i, year in enumerate(full_years):
    row = 5 + i
    recovery = round(conc_total_annotated[i] / proc_total[i] * 100, 1) if proc_total[i] > 0 else 0
    values = [year, conc_cgp3[i], conc_cgp2[i], conc_cgp1[i], conc_tgp[i], conc_total_annotated[i], recovery]
    for col, v in enumerate(values, 1):
        ws4.cell(row=row, column=col, value=v)
        style_data_cell(ws4, row, col)

# Totals
total_row4 = 5 + len(full_years)
ws4.cell(row=total_row4, column=1, value="LoM TOTAL")
ws4.cell(row=total_row4, column=2, value=round(sum(conc_cgp3), 2))
ws4.cell(row=total_row4, column=3, value=round(sum(conc_cgp2), 2))
ws4.cell(row=total_row4, column=4, value=round(sum(conc_cgp1), 2))
ws4.cell(row=total_row4, column=5, value=round(sum(conc_tgp), 2))
ws4.cell(row=total_row4, column=6, value=round(sum(conc_total_annotated), 1))
avg_recovery = round(sum(conc_total_annotated) / sum(proc_total) * 100, 1)
ws4.cell(row=total_row4, column=7, value=avg_recovery)
for col in range(1, len(headers4) + 1):
    style_data_cell(ws4, total_row4, col, is_total_row=True)

auto_width(ws4, len(headers4))

# ──────────────────────────────────────────────────────────────
# SHEET 5: Summary & Analysis
# ──────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Summary & Analysis")

ws5.merge_cells("A1:I1")
ws5["A1"] = "COMPREHENSIVE SUMMARY - Greenbushes LoM Key Metrics (CY25 ORE)"
ws5["A1"].font = TITLE_FONT

ws5.merge_cells("A2:I2")
ws5["A2"] = "All metrics combined for the even-numbered years. Source: Figure 14, Page 28."
ws5["A2"].font = NOTE_FONT

headers5 = [
    "Year",
    "Ore Mined (Mt)", "Waste Mined (Mt)", "Total Movement (Mt)", "Strip Ratio",
    "Stockpile (Mt)",
    "Total Processing (Mt)", "Li2O Grade (%)",
    "Total Concentrate (Mt)"
]
for col, h in enumerate(headers5, 1):
    ws5.cell(row=4, column=col, value=h)
style_header_row(ws5, 4, len(headers5))

for i, year in enumerate(full_years):
    row = 5 + i
    if year in mining_years:
        mi = mining_years.index(year)
        ore_v, waste_v, total_v, sr_v = ore_mt[mi], waste_mt[mi], total_movement[mi], strip_ratios[mi]
    else:
        ore_v, waste_v, total_v, sr_v = 0, 0, 0, "N/A"

    values = [year, ore_v, waste_v, total_v, sr_v,
              stockpiles[i], proc_total[i], li2o_grade[i], conc_total_annotated[i]]
    for col, v in enumerate(values, 1):
        ws5.cell(row=row, column=col, value=v)
        style_data_cell(ws5, row, col)

auto_width(ws5, len(headers5))

# ──────────────────────────────────────────────────────────────
# SHEET 6: Key Insights
# ──────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Key Insights")

ws6.merge_cells("A1:B1")
ws6["A1"] = "KEY INSIGHTS & DEEP ANALYSIS"
ws6["A1"].font = TITLE_FONT

insights = [
    ("LoM Duration", "Mining: 2026-2049 (24 years). Processing continues to ~2053 to exhaust stockpiles."),
    ("Peak Mining Year", f"2034 - Total movement ~{max(total_movement)} Mt with highest strip ratio of {max(strip_ratios)}. This is a massive waste stripping campaign."),
    ("Total Ore Mined (even yrs)", f"{round(sum(ore_mt), 1)} Mt across even-year snapshots"),
    ("Total Waste Mined (even yrs)", f"{round(sum(waste_mt), 1)} Mt across even-year snapshots"),
    ("Average Strip Ratio", f"{round(sum(waste_mt)/sum(ore_mt), 2)} (waste:ore) - Very high waste burden, especially in 2034"),
    ("Peak Stockpile", f"{max(stockpiles)} Mt in {full_years[stockpiles.index(max(stockpiles))]}"),
    ("Stockpile Strategy", "Build-up phase 2026-2032 (peaks 23.4 Mt). Drawn down through mid-LoM. Second smaller peak of 20.7 Mt in 2038. Exhausted by 2053."),
    ("Processing Capacity", "~7-8 Mt/year across 4 plants (CGP3, CGP2, CGP1, TGP) during steady state operations"),
    ("Average Li2O Grade", f"{round(sum(li2o_grade)/len(li2o_grade), 2)}% - Grades decline significantly in tail years (1.19% in 2050, 1.79% in 2052)"),
    ("Total Concentrate (even yrs)", f"{round(sum(conc_total_annotated), 1)} Mt across even-year snapshots"),
    ("Concentrate Recovery Proxy", f"~{avg_recovery}% (concentrate / feed) - relatively stable through mine life"),
    ("Mining Wind-Down", "Strip ratio drops from 11.5 (2034 peak) to 0.8 (2048), transition to lower waste. Less material moved but higher ore proportion."),
    ("2034 Anomaly", "Highest total movement (~69 Mt) but only ~5.5 Mt ore - massive waste stripping campaign with strip ratio 11.5x. This is the most capital-intensive mining year."),
    ("Post-Mining Phase", "2050-2053: Processing from stockpiles only. Grade drops significantly (1.19-1.79% Li2O). Concentrate output collapses to 0.1-0.4 Mt."),
    ("CGP2 Dominance", "CGP2 is the largest processing plant, handling ~2.5-3.0 Mt/year. It contributes the most concentrate output across the LoM."),
    ("CGP3 Ramp-Up", "CGP3 starts small (0.3 Mt in 2026) and ramps to 1.5 Mt by 2032, suggesting a newer plant coming online."),
    ("Production Plateau", "Concentrate production is remarkably stable at 1.6-2.0 Mt/year from 2028-2048, providing consistent spodumene supply."),
    ("Grade Risk", "Li2O grade trends downward from 2.22% peak (2030) toward ~1.95% by 2042-2046. Post-mining grades drop sharply. Lower grades mean higher processing costs per tonne of concentrate."),
]

for col, h in enumerate(["Metric", "Detail"], 1):
    ws6.cell(row=3, column=col, value=h)
style_header_row(ws6, 3, 2)

for i, (metric, detail) in enumerate(insights):
    row = 4 + i
    ws6.cell(row=row, column=1, value=metric)
    ws6.cell(row=row, column=2, value=detail)
    ws6.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
    ws6.cell(row=row, column=2).font = DATA_FONT
    ws6.cell(row=row, column=1).border = THIN_BORDER
    ws6.cell(row=row, column=2).border = THIN_BORDER
    ws6.cell(row=row, column=1).alignment = Alignment(vertical="top")
    ws6.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

ws6.column_dimensions["A"].width = 30
ws6.column_dimensions["B"].width = 100

# ──────────────────────────────────────────────────────────────
# SHEET 7: Data Notes
# ──────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("Data Notes")
ws7["A1"] = "DATA SOURCE & METHODOLOGY NOTES"
ws7["A1"].font = TITLE_FONT

notes_text = [
    "Source: Figure 14, Page 28 - 'Greenbushes by year LoM key metrics for the CY25 ORE' (IGO Limited Annual Report)",
    "",
    "GRAPH 1 - MINING (Mt):",
    "  - Strip ratios directly annotated: 3.2, 3.1, 4.3, 5.8, 11.5, 5.0, 4.5, 5.5, 3.4, 1.9, 1.4, 0.8",
    "  - Total movement estimated from chart y-axis (0-80 Mt scale)",
    "  - Ore = Total / (1 + Strip Ratio); Waste = Total - Ore",
    "  - Mining ceases at end of 2049",
    "",
    "GRAPH 2 - STOCKPILES (Mt):",
    "  - Values directly annotated: 8.2, 17.6, 21.1, 23.4, 17.0, 16.7, 20.7, 16.2, 14.6, 14.2, 11.2, 9.5, 2.7, 0.5",
    "  - Stockpiles are mostly off-RoM",
    "",
    "GRAPH 3 - PROCESSING (Mt):",
    "  - Mean lithia feed grade (%Li2O) directly annotated for all plants combined",
    "  - Plant breakdown (CGP3, CGP2, CGP1, TGP) estimated from stacked area chart bands",
    "  - Processing continues through 2053 to exhaust stockpiles",
    "  - The LoM does not consider underground feed at the production tail",
    "",
    "GRAPH 4 - CONCENTRATE (Mt):",
    "  - Total concentrate values directly annotated: 1.5, 2.0, 2.0, 1.7, 1.8, 1.7, 1.6, 1.6, 1.6, 1.7, 1.6, 1.9, 0.4, 0.1",
    "  - Plant breakdown estimated from stacked area chart bands",
    "",
    "GENERAL NOTES:",
    "  - Data is for even-numbered years only (biannual snapshots starting 2026)",
    "  - These values are the values associated with the even-numbered years, NOT two-year means",
    "  - Plant breakdown for Processing and Concentrate are visual estimates from stacked area charts",
    "  - Annotated values (strip ratios, stockpiles, grades, concentrate totals) are exact as shown on graph",
    "  - Total movement (Mining chart) is estimated from the chart and may have +/- 2-3 Mt margin of error",
]

for i, line in enumerate(notes_text):
    ws7.cell(row=3 + i, column=1, value=line)
    if line.endswith(":"):
        ws7.cell(row=3 + i, column=1).font = SUBTITLE_FONT
    elif line.startswith("  -"):
        ws7.cell(row=3 + i, column=1).font = NOTE_FONT
    else:
        ws7.cell(row=3 + i, column=1).font = DATA_FONT

ws7.column_dimensions["A"].width = 120

# ==============================================================================
# SAVE
# ==============================================================================

output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Greenbushes_LoM_Analysis.xlsx")
wb.save(output_path)
print(f"Excel workbook saved to: {output_path}")
print(f"\nSheets created:")
print(f"  1. Mining - Ore, Waste, Total Movement, Strip Ratio")
print(f"  2. Stockpiles - Stockpile levels with changes")
print(f"  3. Processing - Plant-by-plant processing with Li2O grades")
print(f"  4. Concentrate - Plant-by-plant concentrate output")
print(f"  5. Summary & Analysis - All metrics combined")
print(f"  6. Key Insights - Deep analysis and observations")
print(f"  7. Data Notes - Source and methodology documentation")
